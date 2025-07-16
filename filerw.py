import os
import re

import chardet
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import json


def read(file_path):
    # 先以二进制模式读取文件一小部分，用于检测编码
    with open(file_path, 'rb') as f:
        raw_data = f.read(1000)  # 读取前1000字节数据
        result = chardet.detect(raw_data)  # 使用chardet库检测编码
        encoding = result['encoding']  # 获取检测到的编码

    # 再以检测到的编码打开文件读取全部内容
    with open(file_path, 'r', encoding=encoding) as f:
        content_list = f.readlines()  # 读取文件内容为列表，每个元素是一行文本

    # 去除每行中的异常空格
    cleaned_content_list = []
    for line in content_list:
        # 使用正则表达式去除多余的空格，包括行首行尾空格以及连续的空格
        cleaned_line = re.sub(r'\s+', ' ', line).strip()
        cleaned_content_list.append(cleaned_line)

    return cleaned_content_list


def write2json(file_path, json_str):
    # 检查文件是否存在，如果不存在则初始化为一个空列表
    if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
        existing_data = []
    else:
        # 读取现有JSON文件内容
        with open(file_path, 'r', encoding='utf-8') as file:
            existing_data = json.load(file)

    # 解析传入的JSON字符串
    new_data = json.loads(json_str)

    # 合并数据
    if isinstance(existing_data, list):
        existing_data.append(new_data)
    else:
        print(f"写入失败：{json_str}")

    # 写回文件
    with open(file_path, 'w', encoding='utf-8') as file:
        json.dump(existing_data, file, ensure_ascii=False, indent=4)


def clear_or_create_file(file_path):
    # 获取文件所在的目录路径
    directory = os.path.dirname(file_path)

    # 确保目录存在（如果不存在则创建）
    if not os.path.exists(directory):
        os.makedirs(directory, exist_ok=True)  # exist_ok=True 防止目录已存在时出错

    # 现在可以安全地创建或清空文件
    with open(file_path, 'w', encoding='utf-8') as file:
        file.write('')



def sort_paths(data):
    """
    根据路径的层级进行排序。
    """
    return sorted(data, key=lambda x: x['path'].strip('/').split('/'))


def generate_path_excel(json_file, output_file):
    """
    根据JSON文件生成Excel文件，路径按层级分列，支持动态路径深度。

    :param json_file: 包含路径信息的JSON文件
    :param output_file: 输出的Excel文件名
    """
    # 读取JSON数据
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 对路径进行排序
    data = sort_paths(data)

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Paths"

    # 找到最长路径的层级数
    max_depth = max(len(item['path'].strip("/").split("/")) if item['path'] else 0 for item in data)

    # 设置表头
    headers = ["id", "Domain", "url", "path", "Status", "Length", "Title", "Params"] + [f"第{i}级路径" for i in range(1, max_depth + 1)]
    ws.append(headers)

    # 样式设置
    alignment = Alignment(horizontal="center", vertical="center")
    font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.alignment = alignment
        cell.font = font
        cell.border = thin_border

    # 填充数据
    for idx, item in enumerate(data, start=1):
        path_parts = [f"/{part}" for part in item['path'].strip("/").split("/")] if item['path'] else []
        row = [
            idx,  # 序号
            item['domain'],
            item['url'],
            item['path'],
            item['status'],
            item['length'],
            item['title'],
            item['params']
            *path_parts
        ]
        ws.append(row)

    # 写入数据并设置样式
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = alignment
            cell.border = thin_border

    # 合并单元格并居中
    def merge_and_center(ws, col_start, col_end, data_col):
        """
        合并相同单元格并设置居中对齐
        :param ws: Worksheet对象
        :param col_start: 开始列号
        :param col_end: 结束列号
        :param data_col: 用于判断合并的列索引
        """
        start_row = 2
        for row in range(3, ws.max_row + 1):
            if ws.cell(row=row, column=data_col).value != ws.cell(row=row - 1, column=data_col).value:
                if start_row != row - 1:
                    ws.merge_cells(start_row=start_row, start_column=col_start, end_row=row - 1, end_column=col_end)
                start_row = row
        if start_row != ws.max_row + 1:
            ws.merge_cells(start_row=start_row, start_column=col_start, end_row=ws.max_row, end_column=col_end)

    # 对domain列和路径层级列进行合并单元格
    merge_and_center(ws, 2, 2, 2)  # 合并Domain列
    for col in range(8, 8 + max_depth):  # 路径层级从第8列开始
        merge_and_center(ws, col, col, col)

    # 设置列宽自动调整
    for col in range(1, ws.max_column + 1):
        max_length = max((len(str(ws.cell(row=row, column=col).value or "")) for row in range(1, ws.max_row + 1)), default=10)
        ws.column_dimensions[get_column_letter(col)].width = max_length + 2

    # 保存文件
    wb.save(output_file)
    print(f"Excel文件已保存为：{output_file}")



