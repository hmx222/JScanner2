import re
import os
import time
from typing import List, Dict, Any, Iterable, Optional, Set, Tuple
from urllib.parse import urlparse, urlunparse
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
# from rich import print

# 类型别名
UrlData = Dict[str, str]
InputData = List[Dict[str, Any]]


def sort_by_domain_and_url(data: List[UrlData]) -> List[UrlData]:
    return sorted(data, key=lambda x: (x['domain'], x['url']))


def normalize_url(url: str) -> str:
    try:
        parsed = urlparse(url)
        scheme = parsed.scheme if parsed.scheme else 'http'
        netloc = parsed.netloc if parsed.netloc else (parsed.path.split('/')[0] if parsed.path else '')
        path = parsed.path.rstrip('/') if parsed.path else ''
        normalized = urlunparse((scheme, netloc, path, '', '', ''))
        return normalized.lower()
    except Exception:
        return url.strip().lower()


class SafePathExcelGenerator:
    """
    智能Excel生成器 - ✅ 支持断点续写 (Append Mode)
    - 自动检测文件是否存在，存在则追加，不存在则新建
    - 完美适配多进程接力扫描场景
    """
    HEADERS = ["id", "Domain", "URL", "Path", "sourceURL"]

    def __init__(self, output_file: str):
        self.output_file = output_file

        # 去重集合 (注意：这里的内存去重只针对本次运行的批次，
        # 跨进程去重主要靠 DiskBloomFilter，这里只是 Excel 写入层面的二次保险)
        self.existing_signatures: Set[Tuple[str, str, str]] = set()
        self.existing_urls: Set[str] = set()

        # 样式定义
        self.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        self.header_font = Font(bold=True, name="Arial")
        self.thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # ✅ [核心修改] 检测文件是否存在，实现续写逻辑
        if os.path.exists(self.output_file):
            try:
                print(f"📂 检测到已存在结果文件: {self.output_file}，加载以进行续写...")
                self.wb = load_workbook(self.output_file)
                self.ws = self.wb.active
                # 获取当前最大行数，作为新 ID 的起点
                self.row_count = self.ws.max_row
                print(f"✅ 文件加载成功，当前行数: {self.row_count}")
            except Exception as e:
                print(f"❌ 加载旧文件失败，将创建新文件 (旧文件可能已损坏): {e}")
                self._create_new_workbook()
        else:
            self._create_new_workbook()

    def _create_new_workbook(self):
        """创建全新的工作簿"""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Paths"
        # 写入表头
        self.ws.append(self.HEADERS)
        self.row_count = 1  # 只有新建时才重置为1 (表头占1行)

        # 设置样式
        for col_idx in range(1, len(self.HEADERS) + 1):
            cell = self.ws.cell(row=1, column=col_idx)
            cell.alignment = self.alignment
            cell.font = self.header_font
            cell.border = self.thin_border

        self.set_initial_column_widths()
        self.save()  # 立即保存一次，占坑

    def set_initial_column_widths(self) -> None:
        """预设列宽"""
        column_widths = {
            1: 6, 2: 20, 3: 50, 4: 30, 5: 50
        }
        for col_idx, width in column_widths.items():
            col_letter = get_column_letter(col_idx)
            self.ws.column_dimensions[col_letter].width = width

    @staticmethod
    def clean_illegal_chars(text: Any) -> Any:
        if not isinstance(text, str): return text
        return re.sub(r'[\x00-\x08\x0b-\x1f]', '', text)

    def _extract_domain_from_url(self, url: str) -> str:
        try:
            parsed = urlparse(url)
            return parsed.netloc or (parsed.path.split('/')[0] if parsed.path else '')
        except Exception:
            return ''

    def _extract_path_from_url(self, url: str) -> str:
        try:
            parsed = urlparse(url)
            return parsed.path or ''
        except Exception:
            return ''

    def _process_input_data(self, input_data: InputData) -> List[UrlData]:
        result = []
        for item in input_data:
            if not isinstance(item, dict) or "next_urls" not in item: continue

            next_urls = item.get("next_urls", [])
            source_url = str(item.get("sourceURL", "")).strip()

            if not isinstance(next_urls, (list, set, tuple)) or isinstance(next_urls, str): continue

            for url in next_urls:
                if not isinstance(url, str) or not url.strip(): continue
                url_str = url.strip()

                # 简单过滤 .js .vue，保持和你原逻辑一致
                if ".js" in url_str or ".vue" in url_str: continue

                domain = self._extract_domain_from_url(url_str)
                path = self._extract_path_from_url(url_str)
                normalized_url = normalize_url(url_str)

                # 内存去重 (注意：每次重启进程这里都是空的，所以跨进程重复无法在这里拦截，
                # 但前面的 BloomFilter 已经拦截过了，这里作为保险)
                signature = (domain.lower(), normalized_url, path.lower())
                if signature in self.existing_signatures or normalized_url in self.existing_urls:
                    continue

                result.append({
                    "url": url_str,
                    "domain": domain,
                    "path": path,
                    "sourceURL": source_url
                })
                self.existing_signatures.add(signature)
                self.existing_urls.add(normalized_url)
        return result

    def append_data_batch(self, input_data: InputData, batch_size: int = 500, show_progress: bool = True) -> None:
        if not input_data: return

        # 1. 处理输入数据
        processed_data = self._process_input_data(input_data)
        if not processed_data: return

        # 2. 按域名排序
        sorted_data = sort_by_domain_and_url(processed_data)

        # 3. 写入逻辑
        # 注意：这里不再显示总进度条，因为是追加写入
        # 如果需要显示，可以用简单的 print
        print(f"📊 [Excel追加] 正在写入 {len(sorted_data)} 条新数据...")

        # 批量写入
        current_rows = []
        for item in sorted_data:
            self.row_count += 1
            row_data = [
                str(self.row_count),  # 连续的 ID
                self.clean_illegal_chars(item["domain"]),
                self.clean_illegal_chars(item["url"]),
                self.clean_illegal_chars(item["path"]),
                self.clean_illegal_chars(item["sourceURL"])
            ]
            self.ws.append(row_data)

            # 设置样式（openpyxl append 后需要重新获取行对象来设置样式，稍微有点慢但为了美观）
            # 为了性能，可以每 100 行设置一次或者最后统一设置，但为了实时性，这里逐行设置
            for col_idx in range(1, len(row_data) + 1):
                cell = self.ws.cell(row=self.row_count, column=col_idx)
                cell.alignment = self.alignment
                cell.border = self.thin_border

        # 保存文件
        self.save()
        print(f"✅ [Excel追加] 完成，当前总行数: {self.row_count}")

    def save(self) -> None:
        try:
            self.wb.save(self.output_file)
        except Exception as e:
            print(f"❌ 保存文件失败 (请检查文件是否被占用): {str(e)}")
